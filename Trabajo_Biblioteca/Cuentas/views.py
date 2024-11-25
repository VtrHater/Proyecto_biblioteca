from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.forms import UserCreationForm, AuthenticationForm
from django.contrib.auth.models import User
from django.contrib.auth import login, logout, authenticate
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from .models import Notification, Profile
from .forms import NotificationForm, solicitudesform,estadosSolicitudesform
from .models import Solicitudes, Personal
import mimetypes
import os
from django.shortcuts import redirect
from .forms import ProfileForm, UserForm
from django.contrib import messages
from django.http.response import HttpResponse
from django.views.generic.base import TemplateView
from openpyxl import Workbook
from openpyxl.styles import Alignment,Border,Font,PatternFill,Side
from datetime import datetime


def registrar(request):
    if request.method == "GET":
        return render(request, "base.html",{
        "form": UserCreationForm,
        })
    else:
        if request.POST['password1'] == request.POST['password2']:
            try:
                user = User.objects.create_user(username=request.POST['username'], password=request.POST['password1'])
                user.save()
                login(request, user)
                return redirect("Home")
            except: 
                return render(request, "base.html",{
                "form": UserCreationForm,
                'error': "Nombre de usuario en uso"
                }) 
        else:
            return render(request, "base.html",{
                "form": UserCreationForm,
                'error': "Contraseñas no coinciden"
                }) 

def tareas(request):
    return render(request, "solicitudes.html")

def Home(request):
    return render(request, 'Home.html')

def cerrar_sesion(request):
    logout(request)
    return redirect('Home')

def iniciar_sesion(request):
    if request.method == 'GET':
        return render(request, 'iniciar_sesion.html',{
        'form': AuthenticationForm
        })
    else:
        user = authenticate(request, username=request.POST['username'], password=request.POST['password'])
        if user is None:
            return render(request, 'iniciar_sesion.html',{
            'form': AuthenticationForm,
            'error': "Nombre de cuenta o contraseña incorrectos"
            })
        else:
            login(request,user)
            return redirect('Home')
def solicitud(request):
    data = {
        'form': solicitudesform()
    }

    if request.method == 'POST':
        formulario = solicitudesform(data=request.POST)
        if formulario.is_valid():
            formulario.save()
            data["mensaje"] = "solicitud enviada"
        else:
            data["form"] = formulario

    return render(request, 'solicitudes.html', data)

def soliexistentes(request):
    variable = Solicitudes.objects.values() #values. variable es la info.
    return render(request, 'solicitudes_existentes.html', {'contexto': variable})

def solicitudes_lista(request):
    variable = Solicitudes.objects.all() #all
    return render(request,'solicitudes_existentes.html', {'contexto': variable})


def filtrar_activas(request):
    if 'activas' in request.GET and request.GET['activas'] == 'true':
        var_activas = Solicitudes.objects.exclude(estado='Completada')
    else:
        var_activas = Solicitudes.objects.all()
    return render(request, 'solicitudes_activas.html', {'contexto': var_activas})

def filtrar_solicitudes_usuario(request):
    if request.user.is_authenticated:
        usuario = request.user.username
        soli_usuario= Solicitudes.objects.filter(funcionario=usuario)
    else:
        soli_usuario= Solicitudes.objects.none()
    return render(request,"mis_solicitudes.html", {"contexto":soli_usuario})


def editar_mis_solicitudes(request):
     return render(request, "editar_mis_solicitudes.html")
        
def editar_personal(request):
    if request.user.is_authenticated:
        usuario = request.user.username
        soli_usuario= Solicitudes.objects.filter(funcionario=usuario)
    else:
        soli_usuario= Solicitudes.objects.none()
    return render(request,"editar_personales.html", {"valor":soli_usuario})

def editar_solicitud_personal(request, lol):
    if request.method == "GET":
        soli_citud = get_object_or_404(Solicitudes, pk= lol)
        soli_form = solicitudesform(instance = soli_citud)
        return render(request, "editar_solicitudes_personal.html", {"mostrar":soli_citud, "formulario": soli_form})
    else:
        actualizada = get_object_or_404(Solicitudes, pk=lol)
        formulario_actualizado = solicitudesform(request.POST, instance=actualizada)
        formulario_actualizado.save()
        return redirect("filtrar_solicitudes_usuario")
    
def editar_por_departamento(request):
    if request.user.is_authenticated:
        usuario= request.user.id
        department= Personal.objects.values_list("sector").filter(name=usuario)
        department= department[0][0]
        entregar= Solicitudes.objects.filter(departamento=department).values()
        return render(request, "editar_departamental.html",{"mostrar":entregar})
    
def editar_departamento(request):
    if request.user.is_authenticated:
        usuario= request.user.id
        department= Personal.objects.values_list("sector").filter(name=usuario)
        department= department[0][0]
        entregar= Solicitudes.objects.filter(departamento=department).values()
        return render(request, "editar_solicitud_departamento.html",{"valor":entregar})
    else:
        soli_usuario= Solicitudes.objects.none()
    return render(request,"editar_solicitud_departamento.html", {"valor":soli_usuario})


def editar_estados(request, lol):
    if request.method == "GET":
        soli_citud = get_object_or_404(Solicitudes, pk= lol)
        soli_form = estadosSolicitudesform(instance = soli_citud)
        return render(request, "editar_solicitudes_personal.html", {"mostrar":soli_citud, "formulario": soli_form})
    else:
        actualizada = get_object_or_404(Solicitudes, pk=lol)
        formulario_actualizado = estadosSolicitudesform(request.POST, instance=actualizada)
        formulario_actualizado.save()
        return redirect("filtrar_solicitudes_usuario_estados")
    
    
def filtrar_solicitudes_usuario_estados(request):
    valor= request.GET
    if request.user.is_authenticated:
        user_name = request.user.username  
        departamento = valor.get('departamento') 
        if departamento and departamento != "placeholder":
            var_s_usuario = var_s_usuario.filter(departamento=departamento)
        var_s_usuario = Solicitudes.objects.filter(funcionario=user_name)
    else:
        var_s_usuario = Solicitudes.objects.none()  
    return render(request, 'editar_departamento.html', {'contexto': var_s_usuario,"activas":valor})
    
def prioridades(request):
    soli_MI= Solicitudes.objects.filter(prioridad="Muy Importante")
    soli_I= Solicitudes.objects.filter(prioridad="Importante")
    soli_PI= Solicitudes.objects.filter(prioridad="Poco Importante")
    return render(request,"prioridad.html", {
        "MI":soli_MI,
        "I":soli_I,
        "PI":soli_PI
    })

def prioridades_usuario(request):
    if request.user.is_authenticated:
        usuario = request.user.id  
        department = Personal.objects.filter(name=usuario).values_list("sector", flat=True).first()
        if department is None:
            department = "placeholder"  
        soli_MI = Solicitudes.objects.filter(prioridad="Muy Importante", departamento=department)
        soli_I = Solicitudes.objects.filter(prioridad="Importante", departamento=department)
        soli_PI = Solicitudes.objects.filter(prioridad="Poco Importante", departamento=department)
        
        return render(request, "prioridad_usuario.html", {
            "MI": soli_MI,
            "I": soli_I,
            "PI": soli_PI
        })

def administrador(request):
    return render(request, "administrador.html")

def descargar(request):
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    nom_archivo = "registro.txt"
    filepath = BASE_DIR + "\\" + nom_archivo
    path = open(filepath, "r")
    mime_type, _ = mimetypes.guess_type(filepath)
    response = HttpResponse(path, content_type = mime_type)
    response["content_disposition"] = "attachment; nombre={}".format(nom_archivo)
    return response

@login_required
def redirigir_solicitudes(request):
    if request.user.groups.filter(name='Usuarios_jefes').exists():
        return redirect('soliexistentes')  
    elif request.user.groups.filter(name='Usuarios_general').exists():
        return redirect('filtrar_solicitudes_usuario') 
    else:
        return redirect('Home')  
    
def soli_dep(request):
    if request.user.is_authenticated:  
        usuario= request.user.id
        department= Personal.objects.values_list("sector").filter(name=usuario)
        department= department[0][0]
        entregar= Solicitudes.objects.filter(departamento=department).values()
        return render(request, 'solicitudes_dep.html', {"contexto":entregar})

@login_required
def editar_perfil(request):
    try:
        profile = request.user.profile
    except Profile.DoesNotExist:
        profile = Profile.objects.create(user=request.user)
    if request.method == 'POST':
        user_form = UserForm(request.POST, instance=request.user)
        profile_form = ProfileForm(request.POST, request.FILES, instance=request.user.profile)
        if user_form.is_valid() and profile_form.is_valid():
            user_form.save()
            profile_form.save()
            messages.success(request, '¡Tus datos se han guardado correctamente!')
            return redirect('perfil') 
    else:
        user_form = UserForm(instance=request.user)
        profile_form = ProfileForm(instance=request.user.profile)

    return render(request, 'cuentas/editar_perfil.html', {
        'user_form': user_form,
        'profile_form': profile_form
    })

@login_required
def enviar_notificacion(request):
    if request.method == 'POST':
        form = NotificationForm(request.POST)
        if form.is_valid():
            usuarios = form.cleaned_data['usuarios']  
            message = form.cleaned_data['message']   
            for usuario in usuarios:
                Notification.objects.create(
                    user=usuario,
                    sender=request.user,  
                    message=message
                )
            
    else:
        form = NotificationForm()

    return render(request, 'cuentas/enviar_notificacion.html', {'form': form})

@login_required
def profile_view(request):
    user_notifications = Notification.objects.filter(user=request.user).order_by('-timestamp')
    unread_count = user_notifications.filter(is_read=False).count()
    return render(request, 'cuentas/profile.html', {
        'notifications': user_notifications,
        'unread_count': unread_count
    })

@login_required
def mark_as_read(request):
    notification = Notification.objects.get(id=notification, user=request.user)
    notification.is_read = True
    notification.save()
    return redirect('profile')

class ReportePersonalizadoExcel(TemplateView):
    def get(self,request,*args,**kwargs):
        campo = request.GET.get('campo')
        query = Solicitudes.objects.filter(departamento = campo)
        wb = Workbook()
        controlador = 4
        ws = wb.active
        ws.title = 'Reporte'
        
        
        for q in query:
            fecha_sin_tz = q.fecha.replace(tzinfo=None) if q.fecha else None


            #Crear título en la hoja
            ws['B1'].alignment = Alignment(horizontal = "center",vertical = "center")
            ws['B1'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                     top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
            ws['B1'].fill = PatternFill(start_color= '66FFCC', end_color= '66FFCC', fill_type= "solid")
            ws['B1'].font = Font(name = 'Calibri', size= 12, bold = True)
            ws['B1'] = 'REPORTE PERSONALIZADO EN EXCEL CON DJANGO'

            
            #Cambiar características de las celdas
            ws.merge_cells('B1:L1')

            ws.row_dimensions[1].height = 25 #Altura
            ws.row_dimensions[3].height = 20

            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 30
            ws.column_dimensions['D'].width = 30
            ws.column_dimensions['E'].width = 30
            ws.column_dimensions['F'].width = 30
            ws.column_dimensions['G'].width = 30
            ws.column_dimensions['H'].width = 30
            ws.column_dimensions['I'].width = 30
            ws.column_dimensions['J'].width = 30
            ws.column_dimensions['K'].width = 30
            ws.column_dimensions['L'].width = 30

            #Crear la cabecera
            ws['B3'].alignment = Alignment(horizontal = "center", vertical= "center")
            ws['B3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                     top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
            ws['B3'].fill = PatternFill(start_color= '66CFCC', end_color= '66CFCC', fill_type= "solid")
            ws['B3'].font = Font(name = 'Calibri', size= 10, bold = True)
            ws['B3'] = 'Documento'

            ws['C3'].alignment = Alignment(horizontal = "center", vertical= "center")
            ws['C3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                     top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
            ws['C3'].fill = PatternFill(start_color= '66CFCC', end_color= '66CFCC', fill_type= "solid")
            ws['C3'].font = Font(name = 'Calibri', size= 10, bold = True)
            ws['C3'] = 'Autor'

            ws['D3'].alignment = Alignment(horizontal = "center", vertical= "center")
            ws['D3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                     top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
            ws['D3'].fill = PatternFill(start_color= '66CFCC', end_color= '66CFCC', fill_type= "solid")
            ws['D3'].font = Font(name = 'Calibri', size= 10, bold = True)
            ws['D3'] = 'Fecha de publicación'

            ws['E3'].alignment = Alignment(horizontal = "center", vertical= "center")
            ws['E3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                     top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
            ws['E3'].fill = PatternFill(start_color= '66CFCC', end_color= '66CFCC', fill_type= "solid")
            ws['E3'].font = Font(name = 'Calibri', size= 10, bold = True)
            ws['E3'] = 'N° de sistema'

            ws['F3'].alignment = Alignment(horizontal = "center", vertical= "center")
            ws['F3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                     top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
            ws['F3'].fill = PatternFill(start_color= '66CFCC', end_color= '66CFCC', fill_type= "solid")
            ws['F3'].font = Font(name = 'Calibri', size= 10, bold = True)
            ws['F3'] = 'Ubicación'

            ws['G3'].alignment = Alignment(horizontal = "center", vertical= "center")
            ws['G3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                     top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
            ws['G3'].fill = PatternFill(start_color= '66CFCC', end_color= '66CFCC', fill_type= "solid")
            ws['G3'].font = Font(name = 'Calibri', size= 10, bold = True)
            ws['G3'] = 'Departamento encargado'

            ws['H3'].alignment = Alignment(horizontal = "center", vertical= "center")
            ws['H3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                     top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
            ws['H3'].fill = PatternFill(start_color= '66CFCC', end_color= '66CFCC', fill_type= "solid")
            ws['H3'].font = Font(name = 'Calibri', size= 10, bold = True)
            ws['H3'] = 'Nota'

            ws['I3'].alignment = Alignment(horizontal = "center", vertical= "center")
            ws['I3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                     top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
            ws['I3'].fill = PatternFill(start_color= '66CFCC', end_color= '66CFCC', fill_type= "solid")
            ws['I3'].font = Font(name = 'Calibri', size= 10, bold = True)
            ws['I3'] = 'Fecha de creación'

            ws['J3'].alignment = Alignment(horizontal = "center", vertical= "center")
            ws['J3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                     top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
            ws['J3'].fill = PatternFill(start_color= '66CFCC', end_color= '66CFCC', fill_type= "solid")
            ws['J3'].font = Font(name = 'Calibri', size= 10, bold = True)
            ws['J3'] = 'Plazo'

            ws['K3'].alignment = Alignment(horizontal = "center", vertical= "center")
            ws['K3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                     top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
            ws['K3'].fill = PatternFill(start_color= '66CFCC', end_color= '66CFCC', fill_type= "solid")
            ws['K3'].font = Font(name = 'Calibri', size= 10, bold = True)
            ws['K3'] = 'Funcionario'

            ws['L3'].alignment = Alignment(horizontal = "center", vertical= "center")
            ws['L3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                     top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
            ws['L3'].fill = PatternFill(start_color= '66CFCC', end_color= '66CFCC', fill_type= "solid")
            ws['L3'].font = Font(name = 'Calibri', size= 10, bold = True)
            ws['L3'] = 'Estado'

            #Pintar los datos
            ws.cell(row = controlador, column = 2).alignment = Alignment(horizontal = "center")
            ws.cell(row = controlador, column = 2).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                     top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
            ws.cell(row= controlador, column = 2).font = Font(name = 'Calibri', size = 8)
            ws.cell(row = controlador, column = 2).value= q.documento

            ws.cell(row = controlador, column = 3).alignment = Alignment(horizontal = "center")
            ws.cell(row = controlador, column = 3).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                     top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
            ws.cell(row= controlador, column = 3).font = Font(name = 'Calibri', size = 8)
            ws.cell(row = controlador, column = 3).value= q.autor

            ws.cell(row = controlador, column = 4).alignment = Alignment(horizontal = "center")
            ws.cell(row = controlador, column = 4).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                     top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
            ws.cell(row= controlador, column = 4).font = Font(name = 'Calibri', size = 8)
            ws.cell(row = controlador, column = 4).value= q.fecha_publicación

            ws.cell(row = controlador, column = 5).alignment = Alignment(horizontal = "center")
            ws.cell(row = controlador, column = 5).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                     top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
            ws.cell(row= controlador, column = 5).font = Font(name = 'Calibri', size = 8)
            ws.cell(row = controlador, column = 5).value= q.número_de_sistema

            ws.cell(row = controlador, column = 6).alignment = Alignment(horizontal = "center")
            ws.cell(row = controlador, column = 6).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                     top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
            ws.cell(row= controlador, column = 6).font = Font(name = 'Calibri', size = 8)
            ws.cell(row = controlador, column = 6).value= q.ubicación

            ws.cell(row = controlador, column = 7).alignment = Alignment(horizontal = "center")
            ws.cell(row = controlador, column = 7).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                     top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
            ws.cell(row= controlador, column = 7).font = Font(name = 'Calibri', size = 8)
            ws.cell(row = controlador, column = 7).value= q.departamento

            ws.cell(row = controlador, column = 8).alignment = Alignment(horizontal = "center")
            ws.cell(row = controlador, column = 8).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                     top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
            ws.cell(row= controlador, column = 8).font = Font(name = 'Calibri', size = 8)
            ws.cell(row = controlador, column = 8).value= q.nota

            ws.cell(row = controlador, column = 9).alignment = Alignment(horizontal = "center")
            ws.cell(row = controlador, column = 9).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                     top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
            ws.cell(row= controlador, column = 9).font = Font(name = 'Calibri', size = 8)
            ws.cell(row = controlador, column = 9).value= fecha_sin_tz

            ws.cell(row = controlador, column = 10).alignment = Alignment(horizontal = "center")
            ws.cell(row = controlador, column = 10).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                     top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
            ws.cell(row= controlador, column = 10).font = Font(name = 'Calibri', size = 8)
            ws.cell(row = controlador, column = 10).value= q.plazo

            ws.cell(row = controlador, column = 11).alignment = Alignment(horizontal = "center")
            ws.cell(row = controlador, column = 11).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                     top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
            ws.cell(row= controlador, column = 11).font = Font(name = 'Calibri', size = 8)
            ws.cell(row = controlador, column = 11).value= q.funcionario

            ws.cell(row = controlador, column = 12).alignment = Alignment(horizontal = "center")
            ws.cell(row = controlador, column = 12).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                     top = Side(border_style = "thin"), bottom = Side(border_style = "thin") )
            ws.cell(row= controlador, column = 12).font = Font(name = 'Calibri', size = 8)
            ws.cell(row = controlador, column = 12).value= q.estado
            
            controlador += 1

        #Establecer el nombre de mi archivo
        nombre_archivo = "ReportePersonalizadoExcel.xlsx"
        #Definir el tipo de respuesta que se va a dar
        response = HttpResponse(content_type = "application/ms-excel")
        contenido = "attachment; filename = {0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response
